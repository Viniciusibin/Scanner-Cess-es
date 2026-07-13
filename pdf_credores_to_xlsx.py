from __future__ import annotations

import re
import statistics
from datetime import datetime
from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


COMPANIES = sorted(
    [
        "CRISTOVAO LOCADORA DE VEICULOS",
        "AGRÍCOLA SÃO JUDAS LTDA",
        "TREBESCHI TOMATES MINAS LTDA",
        "TREBESCHI TOMATES CEARA LTDA",
        "TREBESCHI TOMATES GOIAS LTDA",
        "TREBESCHI TOMATES SUL LTDA",
        "S.CRISTOVAO LOC.VEICULOS",
        "EDSON ANTONIO TREBESCHI",
        "AGR. ESP. SANTO",
        "ERICO TREBESCHI",
    ],
    key=len,
    reverse=True,
)

STATE_SET = {
    "AC",
    "AL",
    "AP",
    "AM",
    "BA",
    "CE",
    "DF",
    "ES",
    "GO",
    "MA",
    "MT",
    "MS",
    "MG",
    "PA",
    "PB",
    "PR",
    "PE",
    "PI",
    "RJ",
    "RN",
    "RS",
    "RO",
    "RR",
    "SC",
    "SP",
    "SE",
    "TO",
}

ADDRESS_STARTERS = {
    "RUA",
    "AV",
    "AV.",
    "AVENIDA",
    "RODOVIA",
    "ROD",
    "ROD.",
    "R",
    "PRAÇA",
    "PRACA",
    "AREA",
    "ÁREA",
    "ESTRADA",
    "FAZENDA",
    "TRAVESSA",
    "ALAMEDA",
    "ANEL",
    "QD",
    "QUADRA",
    "CHACARA",
    "CHÁCARA",
}

REGIME_WORDS = {
    "MENSAL",
    "ÚNICO",
    "UNICO",
    "ANUAL",
    "SEMESTRAL",
    "PARCELA",
    "ÚNICA",
    "UNICA",
    "SEMANAL",
}

DOC_RE = re.compile(
    r"\d{3}\.\d{3}\.\d{3}-\d{2}"
    r"|\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}"
    r"|\b\d{11,14}\b"
)
CEP_RE = re.compile(r"^\d{2}\.?\d{3}-?\d{3}$")
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
MONEY_RE = re.compile(r"^\d{1,3}(?:\.\d{3})*,\d{2}$")
LINE_VALUE_RE = re.compile(r"R\$\s*([\d\.]+,\d{2})|([\d\.]+,\d{2})R\$")
POST_TIME_VALUE_RE = re.compile(r"0:00:00\s*(?:R\$\s*)?([\d\.]+,\d{2})\s*R?\$?$")

HEADERS = [
    "Empresa",
    "Nome credor/Razão Social",
    "CPF",
    "Endereço Eletrônico",
    "Endereço",
    "Bairro",
    "Cidade",
    "UF",
    "CEP",
    "Origem/Natureza",
    "Número do Título",
    "Regime de Vencimento",
    "Vencimentos da última parcela",
    "Valor",
    "Pagina",
]


def clean_text(value: str) -> str:
    value = value.replace("\n", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip(" -")


def normalize_cep(value: str) -> str:
    digits = re.sub(r"\D", "", value)
    if len(digits) == 8:
        return f"{digits[:2]}.{digits[2:5]}-{digits[5:]}"
    return clean_text(value)


def parse_brl(value: str) -> float | None:
    text = clean_text(value).replace("R$", "").strip()
    if not text:
        return None
    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def parse_brl_from_line(raw_text: str) -> float | None:
    match = LINE_VALUE_RE.search(raw_text)
    if not match:
        return None
    return parse_brl(match.group(1) or match.group(2) or "")


def parse_brl_after_timestamp(raw_text: str) -> float | None:
    match = POST_TIME_VALUE_RE.search(raw_text)
    if not match:
        return None
    return parse_brl(match.group(1) or "")


def parse_datetime(value: str) -> datetime | None:
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def group_rows(words: list[dict], tolerance: float = 1.0) -> list[list[dict]]:
    rows: list[list[dict]] = []
    row_tops: list[float] = []
    for word in sorted(words, key=lambda item: (item["y0"], item["x0"])):
        if not rows or abs(word["y0"] - row_tops[-1]) > tolerance:
            rows.append([word])
            row_tops.append(word["y0"])
        else:
            rows[-1].append(word)
            row_tops[-1] = (row_tops[-1] * (len(rows[-1]) - 1) + word["y0"]) / len(rows[-1])
    return rows


def words_to_text(words: list[dict]) -> str:
    return clean_text(" ".join(word["text"] for word in sorted(words, key=lambda item: item["x0"])))


def split_company_and_creditor(raw_text: str) -> tuple[str, str]:
    company = ""
    for prefix in COMPANIES:
        if raw_text.startswith(prefix):
            company = prefix
            break

    if not company:
        return "", ""

    doc_match = DOC_RE.search(raw_text)
    if not doc_match:
        return company, clean_text(raw_text[len(company) :])

    creditor = clean_text(raw_text[len(company) : doc_match.start()])
    return company, creditor


def join_words(words: list[dict]) -> str:
    return clean_text(" ".join(word["text"] for word in sorted(words, key=lambda item: item["x0"])))


def split_by_largest_gap(
    words: list[dict],
    min_gap: float,
    fallback: str,
) -> tuple[list[dict], list[dict]]:
    if not words:
        return [], []
    if len(words) == 1:
        return (words, []) if fallback == "left" else ([], words)

    best_index = -1
    best_gap = -1.0
    ordered = sorted(words, key=lambda item: item["x0"])
    for index in range(len(ordered) - 1):
        gap = ordered[index + 1]["x0"] - ordered[index]["x1"]
        if gap > best_gap:
            best_gap = gap
            best_index = index

    if best_gap >= min_gap:
        return ordered[: best_index + 1], ordered[best_index + 1 :]

    return (ordered, []) if fallback == "left" else ([], ordered)


def infer_layout(page_words: list[dict]) -> dict[str, float]:
    header_map = {}
    for word in page_words:
        if word["text"] in {"Cidade", "UF", "Origem/Natureza", "Regime"}:
            header_map[word["text"]] = word["x0"]

    email_xs = [word["x0"] for word in page_words if "@" in word["text"]]
    address_xs = [word["x0"] for word in page_words if word["text"].upper() in ADDRESS_STARTERS]
    uf_xs = [word["x0"] for word in page_words if word["text"] in STATE_SET]
    cep_xs = [word["x0"] for word in page_words if CEP_RE.fullmatch(word["text"])]
    regime_xs = [word["x0"] for word in page_words if word["text"].upper() in REGIME_WORDS]
    date_xs = [word["x0"] for word in page_words if DATE_RE.fullmatch(word["text"])]
    value_xs = [word["x0"] for word in page_words if word["text"] == "R$"]

    email_x = min(email_xs) if email_xs else 280.0
    address_x = min(address_xs) if address_xs else 340.0
    uf_x = statistics.median(uf_xs) if uf_xs else 607.0
    cep_x = statistics.median(cep_xs) if cep_xs else 619.0
    header_city_x = header_map.get("Cidade", uf_x - 30.0)
    header_uf_x = header_map.get("UF", uf_x)
    header_origin_x = header_map.get("Origem/Natureza", cep_x + 25.0)
    header_regime_x = header_map.get("Regime", 707.0)
    regime_x = min(regime_xs) if regime_xs else header_regime_x
    venc_x = min(date_xs) if date_xs else 733.0
    value_x = min(value_xs) if value_xs else 768.0

    return {
        "email_x": email_x,
        "address_x": address_x,
        "city_x": header_city_x - ((header_uf_x - header_city_x) / 2.0) - 4.0,
        "uf_x": uf_x,
        "cep_x": cep_x,
        "origin_x": (cep_x + header_origin_x) / 2.0,
        "regime_x": regime_x,
        "venc_x": venc_x,
        "value_x": value_x,
    }


def extract_rows(pdf_path: Path) -> list[dict]:
    document = fitz.open(str(pdf_path))
    rows_out: list[dict] = []
    previous_layout: dict[str, float] | None = None

    for page_number, page in enumerate(document, start=1):
        words = []
        for x0, y0, x1, y1, text, *_ in page.get_text("words"):
            if text.strip():
                words.append(
                    {
                        "x0": float(x0),
                        "y0": float(y0),
                        "x1": float(x1),
                        "y1": float(y1),
                        "text": text.strip(),
                    }
                )

        if not words:
            continue

        header_hits = sum(
            1 for word in words if word["text"] in {"Cidade", "UF", "CEP", "Origem/Natureza", "Valor"}
        )
        if previous_layout is not None and header_hits < 3:
            layout = previous_layout
        else:
            layout = infer_layout(words)
            previous_layout = layout

        page_rows = group_rows(words)

        for row_words in page_rows:
            raw_text = words_to_text(row_words)
            if "0:00:00" not in raw_text or not re.search(r"\d{1,3}(?:\.\d{3})*,\d{2}", raw_text):
                continue

            ordered = sorted(row_words, key=lambda item: item["x0"])
            company, creditor = split_company_and_creditor(raw_text)
            doc_match = DOC_RE.search(raw_text)
            document_id = doc_match.group(0) if doc_match else ""

            email_words = [w for w in ordered if layout["email_x"] <= w["x0"] < layout["address_x"]]
            geo_words = [w for w in ordered if layout["address_x"] <= w["x0"] < layout["origin_x"]]

            uf_index = next((i for i, w in enumerate(geo_words) if w["text"] in STATE_SET), None)
            cep_index = None
            if uf_index is not None:
                for i in range(uf_index + 1, len(geo_words)):
                    if CEP_RE.fullmatch(geo_words[i]["text"]):
                        cep_index = i
                        break

            if uf_index is not None:
                before_uf = geo_words[:uf_index]
                city_words = [w for w in before_uf if w["x0"] >= layout["city_x"]]
                local_words = [w for w in before_uf if w["x0"] < layout["city_x"]]
                uf_words = [geo_words[uf_index]]
                cep_words = [geo_words[cep_index]] if cep_index is not None else []
                tail_start = cep_index + 1 if cep_index is not None else uf_index + 1
                tail_words = geo_words[tail_start:] + [
                    w for w in ordered if layout["origin_x"] <= w["x0"] < layout["regime_x"]
                ]
            else:
                local_words = [w for w in ordered if layout["address_x"] <= w["x0"] < layout["city_x"]]
                city_words = [w for w in ordered if layout["city_x"] <= w["x0"] < layout["uf_x"]]
                uf_words = [w for w in ordered if layout["uf_x"] <= w["x0"] < layout["cep_x"]]
                cep_words = [w for w in ordered if layout["cep_x"] <= w["x0"] < layout["origin_x"]]
                tail_words = [w for w in ordered if layout["origin_x"] <= w["x0"] < layout["regime_x"]]

            regime_words = [w for w in ordered if layout["regime_x"] <= w["x0"] < layout["venc_x"]]
            venc_words = [w for w in ordered if layout["venc_x"] <= w["x0"] < layout["value_x"]]
            value_words = [w for w in ordered if w["x0"] >= layout["value_x"]]

            endereco_words, bairro_words = split_by_largest_gap(local_words, min_gap=18.0, fallback="left")
            origem_words, titulo_words = split_by_largest_gap(tail_words, min_gap=14.0, fallback="right")

            email = join_words(email_words)
            endereco = join_words(endereco_words)
            bairro = join_words(bairro_words)
            cidade = join_words(city_words)
            uf = join_words(uf_words)
            cep = normalize_cep(join_words(cep_words))
            origem = join_words(origem_words)
            titulo = join_words(titulo_words)
            regime = join_words(regime_words)
            vencimento_text = join_words(venc_words)
            valor_text = join_words(value_words)
            valor = parse_brl_after_timestamp(raw_text)
            if valor is None:
                valor = parse_brl(valor_text)
            if valor is None:
                valor = parse_brl_from_line(raw_text)

            rows_out.append(
                {
                    "Empresa": company,
                    "Nome credor/Razão Social": creditor,
                    "CPF": document_id,
                    "Endereço Eletrônico": email,
                    "Endereço": endereco,
                    "Bairro": bairro,
                    "Cidade": cidade,
                    "UF": uf,
                    "CEP": cep,
                    "Origem/Natureza": origem,
                    "Número do Título": titulo,
                    "Regime de Vencimento": regime,
                    "Vencimentos da última parcela": parse_datetime(vencimento_text) or vencimento_text,
                    "Valor": valor if valor is not None else valor_text,
                    "Pagina": page_number,
                    "_page": page_number,
                    "_raw": raw_text,
                }
            )

    document.close()
    return rows_out


def autosize(worksheet) -> None:
    for index, header in enumerate(HEADERS, start=1):
        max_len = len(header)
        for row in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row=row, column=index).value
            text = "" if value is None else str(value)
            max_len = max(max_len, len(text))
        worksheet.column_dimensions[get_column_letter(index)].width = min(max_len + 2, 45)


def write_xlsx(rows: list[dict], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Credores"
    sheet.append(HEADERS)

    for row in rows:
        sheet.append([row[header] for header in HEADERS])

    bold_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = bold_font

    for row_index in range(2, sheet.max_row + 1):
        date_cell = sheet.cell(row=row_index, column=13)
        if isinstance(date_cell.value, datetime):
            date_cell.number_format = "dd/mm/yyyy"

        value_cell = sheet.cell(row=row_index, column=14)
        if isinstance(value_cell.value, (int, float)):
            value_cell.number_format = 'R$ #,##0.00'

    autosize(sheet)
    workbook.save(output_path)


def main() -> None:
    pdf_path = next(Path(".").glob("*.pdf"))
    output_path = Path("grupo_trebeschi_com_pagina.xlsx")
    rows = extract_rows(pdf_path)
    write_xlsx(rows, output_path)
    print(f"PDF: {pdf_path.name}")
    print(f"Linhas exportadas: {len(rows)}")
    print(f"Arquivo gerado: {output_path.name}")


if __name__ == "__main__":
    main()
